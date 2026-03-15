from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from datetime import date, datetime, timedelta
from typing import List
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.modules.sessions.models import Session as BookingSession, SessionStatus
from app.modules.masters.models import Professional
from app.modules.services.models import Service
from app.modules.salons.models import Provider
from app.modules.reports.schemas import (
    ProfessionalEarningsSummary, ProviderRevenueSummary, ServicePopularity,
    ProfessionalPerformance, DailyRevenue, ProviderReportResponse, ProfessionalReportResponse,
)


def _date_to_dt(d: date) -> datetime:
    return datetime.combine(d, datetime.min.time())


def get_professional_report(db: Session, professional_id: int, date_from: date, date_to: date) -> ProfessionalReportResponse:
    professional = db.query(Professional).filter(Professional.id == professional_id).first()
    if not professional:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Professional not found")

    sessions = (
        db.query(BookingSession)
        .filter(
            BookingSession.professional_id == professional_id,
            BookingSession.status == SessionStatus.COMPLETED,
            BookingSession.starts_at >= _date_to_dt(date_from),
            BookingSession.starts_at <= _date_to_dt(date_to) + timedelta(days=1),
        )
        .all()
    )

    total_earnings = sum(s.earnings_amount or 0 for s in sessions)
    summary = ProfessionalEarningsSummary(
        professional_id=professional_id,
        professional_name=professional.name,
        sessions_completed=len(sessions),
        total_earnings=total_earnings,
        period_start=date_from,
        period_end=date_to,
    )

    # Daily breakdown
    daily: dict = {}
    for s in sessions:
        day = s.starts_at.date()
        if day not in daily:
            daily[day] = DailyRevenue(date=day, revenue=0, session_count=0)
        daily[day].revenue += s.earnings_amount or 0
        daily[day].session_count += 1

    return ProfessionalReportResponse(
        summary=summary,
        daily_earnings=sorted(daily.values(), key=lambda x: x.date),
    )


# Backward-compat alias
get_master_report = get_professional_report


def get_provider_report(db: Session, provider_id: int, date_from: date, date_to: date) -> ProviderReportResponse:
    provider = db.query(Provider).filter(Provider.id == provider_id).first()
    if not provider:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Provider not found")

    dt_from = _date_to_dt(date_from)
    dt_to = _date_to_dt(date_to) + timedelta(days=1)

    all_sessions = (
        db.query(BookingSession)
        .filter(
            BookingSession.provider_id == provider_id,
            BookingSession.starts_at >= dt_from,
            BookingSession.starts_at < dt_to,
        )
        .all()
    )

    completed = [s for s in all_sessions if s.status == SessionStatus.COMPLETED]
    cancelled = [s for s in all_sessions if s.status == SessionStatus.CANCELLED]
    total_revenue = sum(s.total_paid or 0 for s in completed)
    total_deposits = sum(s.deposit_paid or 0 for s in all_sessions)

    summary = ProviderRevenueSummary(
        provider_id=provider_id,
        provider_name=provider.name,
        period_start=date_from,
        period_end=date_to,
        total_sessions=len(all_sessions),
        completed_sessions=len(completed),
        cancelled_sessions=len(cancelled),
        total_revenue=total_revenue,
        total_deposits=total_deposits,
    )

    # Service popularity
    service_counts: dict = {}
    for s in all_sessions:
        if s.service_id:
            if s.service_id not in service_counts:
                svc = db.query(Service).filter(Service.id == s.service_id).first()
                service_counts[s.service_id] = {"name": svc.name if svc else "Unknown", "count": 0, "revenue": 0.0}
            service_counts[s.service_id]["count"] += 1
            if s.status == SessionStatus.COMPLETED:
                service_counts[s.service_id]["revenue"] += s.total_paid or 0

    service_popularity = [
        ServicePopularity(
            service_id=sid,
            service_name=v["name"],
            booking_count=v["count"],
            total_revenue=v["revenue"],
        )
        for sid, v in sorted(service_counts.items(), key=lambda x: -x[1]["count"])
    ]

    # Professional performance
    professional_stats: dict = {}
    for s in all_sessions:
        if s.professional_id:
            if s.professional_id not in professional_stats:
                p = db.query(Professional).filter(Professional.id == s.professional_id).first()
                professional_stats[s.professional_id] = {"name": p.name if p else "Unknown", "completed": 0, "cancelled": 0, "earnings": 0.0}
            if s.status == SessionStatus.COMPLETED:
                professional_stats[s.professional_id]["completed"] += 1
                professional_stats[s.professional_id]["earnings"] += s.earnings_amount or 0
            elif s.status == SessionStatus.CANCELLED:
                professional_stats[s.professional_id]["cancelled"] += 1

    professional_performance = [
        ProfessionalPerformance(
            professional_id=pid,
            professional_name=v["name"],
            sessions_completed=v["completed"],
            sessions_cancelled=v["cancelled"],
            completion_rate=(
                v["completed"] / max(v["completed"] + v["cancelled"], 1) * 100
            ),
            total_earnings=v["earnings"],
        )
        for pid, v in professional_stats.items()
    ]

    # Daily revenue
    daily: dict = {}
    for s in completed:
        day = s.starts_at.date()
        if day not in daily:
            daily[day] = DailyRevenue(date=day, revenue=0, session_count=0)
        daily[day].revenue += s.total_paid or 0
        daily[day].session_count += 1

    return ProviderReportResponse(
        summary=summary,
        service_popularity=service_popularity,
        professional_performance=professional_performance,
        daily_revenue=sorted(daily.values(), key=lambda x: x.date),
    )


# Backward-compat alias
get_salon_report = get_provider_report


# ── helpers ──────────────────────────────────────────────────────────────────

_PINK_FILL = PatternFill("solid", fgColor="FCE7F3")
_HEADER_FILL = PatternFill("solid", fgColor="831843")
_ALT_FILL = PatternFill("solid", fgColor="FDF2F8")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, color="831843", size=13)
_CURRENCY_FMT = '"$"#,##0.00'


def _header_row(ws, cols: list[str], row: int = 1) -> None:
    for col_idx, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 20


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export_provider_report_xlsx(report: ProviderReportResponse) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Provider Report: {report.summary.provider_name}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Period: {report.summary.period_start} — {report.summary.period_end}"

    _header_row(ws, ["Metric", "Value"], row=4)
    metrics = [
        ("Total Sessions", report.summary.total_sessions),
        ("Completed Sessions", report.summary.completed_sessions),
        ("Cancelled Sessions", report.summary.cancelled_sessions),
        ("Total Revenue", report.summary.total_revenue),
        ("Total Deposits", report.summary.total_deposits),
    ]
    for i, (label, val) in enumerate(metrics, 5):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=val)
        if "Revenue" in label or "Deposits" in label:
            cell.number_format = _CURRENCY_FMT
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = _ALT_FILL
            cell.fill = _ALT_FILL
    _autofit(ws)

    # Sheet 2 — Daily Revenue
    ws2 = wb.create_sheet("Daily Revenue")
    _header_row(ws2, ["Date", "Sessions", "Revenue"])
    for i, day in enumerate(report.daily_revenue, 2):
        ws2.cell(row=i, column=1, value=str(day.date))
        ws2.cell(row=i, column=2, value=day.session_count)
        rev_cell = ws2.cell(row=i, column=3, value=day.revenue)
        rev_cell.number_format = _CURRENCY_FMT
        if i % 2 == 0:
            for c in range(1, 4):
                ws2.cell(row=i, column=c).fill = _ALT_FILL
    _autofit(ws2)

    # Sheet 3 — Professionals
    ws3 = wb.create_sheet("Professionals")
    _header_row(ws3, ["Professional", "Completed", "Cancelled", "Completion %", "Earnings"])
    for i, p in enumerate(report.professional_performance, 2):
        ws3.cell(row=i, column=1, value=p.professional_name)
        ws3.cell(row=i, column=2, value=p.sessions_completed)
        ws3.cell(row=i, column=3, value=p.sessions_cancelled)
        pct = ws3.cell(row=i, column=4, value=p.completion_rate / 100)
        pct.number_format = "0.0%"
        earn = ws3.cell(row=i, column=5, value=p.total_earnings)
        earn.number_format = _CURRENCY_FMT
        if i % 2 == 0:
            for c in range(1, 6):
                ws3.cell(row=i, column=c).fill = _ALT_FILL
    _autofit(ws3)

    # Sheet 4 — Services
    ws4 = wb.create_sheet("Services")
    _header_row(ws4, ["Service", "Bookings", "Revenue"])
    for i, svc in enumerate(report.service_popularity, 2):
        ws4.cell(row=i, column=1, value=svc.service_name)
        ws4.cell(row=i, column=2, value=svc.booking_count)
        rev = ws4.cell(row=i, column=3, value=svc.total_revenue)
        rev.number_format = _CURRENCY_FMT
        if i % 2 == 0:
            for c in range(1, 4):
                ws4.cell(row=i, column=c).fill = _ALT_FILL
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Backward-compat alias
export_salon_report_xlsx = export_provider_report_xlsx


def export_professional_report_xlsx(report: ProfessionalReportResponse) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Professional Report: {report.summary.professional_name}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Period: {report.summary.period_start} — {report.summary.period_end}"

    _header_row(ws, ["Metric", "Value"], row=4)
    metrics = [
        ("Sessions Completed", report.summary.sessions_completed),
        ("Total Earnings", report.summary.total_earnings),
    ]
    for i, (label, val) in enumerate(metrics, 5):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=val)
        if "Earnings" in label:
            cell.number_format = _CURRENCY_FMT
    _autofit(ws)

    # Sheet 2 — Daily Earnings
    ws2 = wb.create_sheet("Daily Earnings")
    _header_row(ws2, ["Date", "Sessions", "Earnings"])
    for i, day in enumerate(report.daily_earnings, 2):
        ws2.cell(row=i, column=1, value=str(day.date))
        ws2.cell(row=i, column=2, value=day.session_count)
        earn = ws2.cell(row=i, column=3, value=day.revenue)
        earn.number_format = _CURRENCY_FMT
        if i % 2 == 0:
            for c in range(1, 4):
                ws2.cell(row=i, column=c).fill = _ALT_FILL
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Backward-compat alias
export_master_report_xlsx = export_professional_report_xlsx
